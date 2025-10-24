from django.shortcuts import render, get_object_or_404, redirect
from .models import Book,Holiday, LibrarySetting,CourseNew, Department, FooterContent, Notification, StudentQuery, IssuedBook, Genre, BookRequestQueue, StudentExtra
from django.contrib.auth.decorators import login_required,user_passes_test
from datetime import datetime, timedelta, date
from django.views.decorators.http import require_POST,  require_GET
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib.auth.decorators import user_passes_test, login_required
from .forms import StudentSignupForm,HolidayForm,DepartmentForm,CourseDepartmentForm,BrandingForm, ExpiryFineForm, AdminCredentialForm,CourseDepartmentForm,EditStudentExtraForm,FooterContentForm,EditStudentExtraForm, EditStudentForm, BookForm, ReplyForm, IssueBookForm, GenreForm, StudentQueryForm
import openpyxl
from decimal import Decimal
from openpyxl import Workbook
from django.utils.dateparse import parse_date
from django.contrib.admin.views.decorators import staff_member_required
from django.http import JsonResponse
from django.contrib.auth import update_session_auth_hash
from django.contrib import messages
from django.utils import timezone
from django.contrib.auth.hashers import make_password
from django.db.models import Q, Count
from django.urls import reverse
from collections import defaultdict
from django.core.paginator import Paginator
from django.utils.dateparse import parse_date
from django.utils.timezone import now
from .utils import send_book_notification_email
from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from django.core.mail import send_mail
import random
import json
from django.core.cache import cache
from django.views.decorators.csrf import csrf_exempt
from twilio.rest import Client
from django.conf import settings
from django.core.paginator import Paginator
from django.db.models import Q 
from django.contrib.auth.views import PasswordResetView, PasswordResetConfirmView
from django.contrib.messages.views import SuccessMessageMixin
from django.contrib import messages
from django.urls import reverse_lazy
from django.shortcuts import render
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_protect
# from .generate_qr_code_view import generate_book_qr
import qrcode
import base64
from io import BytesIO
from django.http import HttpResponse
# from .utils import promote_users 
import barcode
from barcode import Code128
from barcode.writer import ImageWriter
from django.core.files.base import ContentFile
# from django.conf import settings

# views.py
def home(request):
    unset_old_newly_published_books()
    genres = Genre.objects.all()
    genre_books = {genre: Book.objects.filter(genre=genre) for genre in genres}
    bestsellers = Book.objects.filter(bestseller=True)
    today = timezone.now().date()
    one_month_ago = today - timedelta(days=30)
    newly_published = Book.objects.filter(
        newly_published=True,
        published_date__isnull=False,
        published_date__gte=one_month_ago
    )
   
     # Auto-unset books older than 30 days
    Book.objects.filter(
        newly_published=True,
        published_date__lt=one_month_ago
    ).update(newly_published=False)

    return render(request, 'library/index.html', {
        'genres': genres,
        'genre_books': genre_books,
        'bestsellers': bestsellers,
        'newly_published': newly_published,
    })

@staff_member_required
def manage_footer(request):
    footer, created = FooterContent.objects.get_or_create(id=1)

    if request.method == 'POST':
        form = FooterContentForm(request.POST, instance=footer)
        if form.is_valid():
            form.save()
            messages.success(request, "Footer updated successfully!")
            return redirect('manage_footer')  # reload page after save
    else:
        form = FooterContentForm(instance=footer)

    return render(request, 'library/admin_manage_footer.html', {'form': form})

@login_required
def edit_library_settings(request):
    setting, created = LibrarySetting.objects.get_or_create(pk=1)

    branding_form = BrandingForm(request.POST or None, request.FILES or None, instance=setting)
    expiry_form = ExpiryFineForm(request.POST or None, instance=setting)
    credential_form = AdminCredentialForm()

    if request.method == 'POST':
        if 'save_branding' in request.POST:
            if branding_form.is_valid():
                branding_form.save()
                messages.success(request, "Branding updated.")
                return redirect('edit_library_settings')

        elif 'save_expiry_fine' in request.POST:
            if expiry_form.is_valid():
                expiry_form.save()
                messages.success(request, "Expiry and Fine settings updated.")
                return redirect('edit_library_settings')

        elif 'save_credentials' in request.POST:
            credential_form = AdminCredentialForm(request.POST)
            if credential_form.is_valid():
                credential_form.save(request.user)
                messages.success(request, "Admin credentials updated. Please log in again.")
                return redirect('login')  # log out or force re-login

    context = {
        'branding_form': branding_form,
        'expiry_form': expiry_form,
        'credential_form': credential_form,
        'setting': setting,
    }
    return render(request, 'library/edit_settings.html', context)


# published book after 30days
def unset_old_newly_published_books():
    today = timezone.now().date()
    one_month_ago = today - timedelta(days=30)
    Book.objects.filter(
        newly_published=True,
        published_date__lt=one_month_ago
    ).update(newly_published=False)


def search_suggestions(request):
    query = request.GET.get('q', '')
    books = Book.objects.filter(name__icontains=query)[:10]

    results = []
    for book in books:
        borrowed_count = book.issuedbook_set.filter(returned=False, status='Borrowed').count()
        available = book.quantity - borrowed_count
        status = "Available" if available > 0 else "Not available"

        results.append({
            'id': book.id,
            'name': book.name,
            'genre_id': book.genre.id,
            'status': status,
        })

    return JsonResponse({'results': results})

from django.db.models import Sum

def book_detail(request, genre_id, pk):
    genre = get_object_or_404(Genre, id=genre_id)
    book = get_object_or_404(Book, pk=pk, genre_id=genre_id)

    # 🔑 Calculate total quantity across all copies of this book (same title/ISBN)
    total_quantity = Book.objects.filter(name__iexact=book.name).aggregate(Sum('quantity'))['quantity__sum'] or 0

    # Issued/Reserved counts should also consider all copies of this title
    issued_qs = IssuedBook.objects.filter(book__name__iexact=book.name, returned=False).order_by('issued_date')

    borrowed_count = issued_qs.filter(status='Borrowed').count()
    reserved_count = issued_qs.filter(status='Reserved').count()
    available = total_quantity - borrowed_count

    # Get earliest return date
    next_return_date = None
    if available <= 0:
        next_due = issued_qs.filter(status='Borrowed').order_by('expiry_date').first()
        if next_due:
            next_return_date = next_due.expiry_date

    # Already requested?
    already_requested = False
    if request.user.is_authenticated and not request.user.is_superuser:
        already_requested = issued_qs.filter(student=request.user, returned=False).exists()

    # Waiting position
    waiting_position_qs = issued_qs.filter(status='Waiting').values_list('student_id', flat=True)
    days_to_wait = 0
    if available <= 0 and request.user.id in waiting_position_qs:
        pos = list(waiting_position_qs).index(request.user.id)
        days_to_wait = (pos + 1) * 7

    context = {
        'book': book,
        'total_quantity': total_quantity,
        'borrowed_count': borrowed_count,
        'reserved_count': reserved_count,
        'available': available,
        'next_return_date': next_return_date,
        'already_requested': already_requested,
        'days_to_wait': days_to_wait,
    }
    return render(request, 'library/book_detail.html', context)

# ✅ Check if user is superuser
@staff_member_required
def is_admin(user):
    return user.is_authenticated and user.is_superuser




# 🔐 Login view
@staff_member_required
def admin_login(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(username=username, password=password)
        if user is not None and user.is_superuser:
            login(request, user)
            return redirect('admin_dashboard')
        else:
            return render(request, 'library/login.html', {'error': 'Invalid admin credentials'})
    return render(request, 'library/login.html')




@staff_member_required
def is_admin(user):
    return user.is_superuser

@staff_member_required
def admin_dashboard(request):
    today = date.today()
    context = {
        'today_books_count': Book.objects.filter(added_date=today).count(),
        'today_issued_count': IssuedBook.objects.filter(issued_date=today).count(),
        'today_students_count': User.objects.filter(date_joined__date=today).count(),

        'total_books_count': Book.objects.count(),
        'total_students_count': User.objects.count(),
        'total_issued_count': IssuedBook.objects.filter(status='Borrowed').count(),
        'total_reserved_count': IssuedBook.objects.filter(status='Reserved').count(),
        'total_waiting_count': IssuedBook.objects.filter(status='Waiting').count(),
        'total_returned_count': IssuedBook.objects.filter(returned=True).count(),
    }
    return render(request, 'library/admin_dashboard.html', context)



# 🚪 Logout
def admin_logout(request):
    logout(request)
    return redirect('admin_login')

# Check if user is not admin
def is_student(user):
    return user.is_authenticated and not user.is_superuser

@staff_member_required
def manage_course_department(request):
    query = request.GET.get("q", "").strip()  # Search query
    form = CourseDepartmentForm()

    # Base queryset
    departments = Department.objects.select_related("course").all()

    # Apply search filter if query exists
    if query:
        departments = departments.filter(
            Q(name__icontains=query) | 
            Q(course__name__icontains=query)
        )

    #  Export Excel
    if request.GET.get("export") == "xlsx":
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="departments.xlsx"'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Departments"
        ws.append(["Course", "Department"])
        for dept in departments:
            ws.append([dept.course.name, dept.name])
        wb.save(response)
        return response

    # Handle POST (Add course + department)
    if request.method == "POST":
        form = CourseDepartmentForm(request.POST)
        if form.is_valid():
            course_instance = form.cleaned_data.get("course")
            new_course_name = form.cleaned_data.get("new_course")
            department_name = form.cleaned_data.get("department")

            if new_course_name:
                course_instance, created = CourseNew.objects.get_or_create(name=new_course_name)
                if created:
                    messages.success(request, f"New course '{new_course_name}' was created.")
                else:
                    messages.info(request, f"Course '{new_course_name}' already exists.")

            if not course_instance:
                form.add_error("course", "Please select an existing course or enter a new one.")
                messages.error(request, "You must select or create a course.")
                return render(request, "library/manage_course_department.html", {
                    "form": form,
                    "departments": departments,
                })

            if department_name:
                department, created = Department.objects.get_or_create(course=course_instance, name=department_name)
                if created:
                    messages.success(request, f"Department '{department_name}' was added to course '{course_instance.name}'.")
                else:
                    messages.warning(request, f"Department '{department_name}' already exists for course '{course_instance.name}'.")
            else:
                form.add_error("department", "Please provide a department name.")
                messages.error(request, "You must provide a department name.")
                return render(request, "library/manage_course_department.html", {
                    "form": form,
                    "departments": departments,
                })

            return redirect("manage_course_department")

    return render(
        request,
        "library/manage_course_department.html",
        {
            "form": form,
            "courses": CourseNew.objects.all(),
            "departments": departments,  # filtered list
        },
    )



def edit_course_department(request, pk):
    dept = get_object_or_404(Department, pk=pk)
    current_course = dept.course  

    if request.method == "POST":
        form = CourseDepartmentForm(request.POST)

        if form.is_valid():
            course = form.cleaned_data.get("course")
            new_course_name = form.cleaned_data.get("new_course")
            department_name = form.cleaned_data.get("department")

            # ✅ Rename course safely
            if new_course_name:
                current_course.name = new_course_name
                current_course.save()

            # ✅ Reassign department to selected course
            if course:
                dept.course = course  

            # ✅ Update department name
            if department_name:
                dept.name = department_name  

            dept.save()
            messages.success(request, "✅ Course & Department updated successfully.")
            return redirect("manage_course_department")
    else:
        form = CourseDepartmentForm(initial={
            "course": current_course,
            "department": dept.name,
        })

    return render(request, "library/edit_course_department.html", {
        "form": form,
        "dept": dept,
    })

@staff_member_required
def delete_course(request, course_id):
    course = get_object_or_404(CourseNew, id=course_id)
    course.delete()
    messages.success(request, "Course deleted successfully.")
    return redirect('manage_course_department')

@staff_member_required
def delete_department(request, dept_id):
    dept = get_object_or_404(Department, id=dept_id)
    dept.delete()
    messages.success(request, "Department deleted successfully.")
    return redirect('manage_course_department')


from django.contrib.auth.models import User
from django.shortcuts import render, redirect
from django.contrib import messages
from django.core.serializers.json import DjangoJSONEncoder
import json
from .models import StudentExtra, CourseNew, Department
from .forms import StudentSignupForm
from django.contrib.admin.views.decorators import staff_member_required


@staff_member_required
def student_signup(request):
    # Fetch courses & related departments for dropdown
    courses = CourseNew.objects.prefetch_related('departments').all()
    department_map = {
        course.id: [{'id': dept.id, 'name': dept.name} for dept in course.departments.all()]
        for course in courses
    }

    if request.method == 'POST':
        form = StudentSignupForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            email = form.cleaned_data['email']
            mobile_number = form.cleaned_data['mobile_number']
            roll_number = form.cleaned_data['roll_number']
            course = form.cleaned_data.get('course')          # can be None
            department = form.cleaned_data.get('department')  # can be None
            year = form.cleaned_data['year']
            password1 = form.cleaned_data['password1']

            # Duplicate checks
            if User.objects.filter(username=username).exists():
                form.add_error('username', "⚠️ Username already taken.")
            elif User.objects.filter(email=email).exists():
                form.add_error('email', "⚠️ Email already in use.")
            elif StudentExtra.objects.filter(roll_number=roll_number).exists():
                form.add_error('roll_number', "⚠️ Roll number already exists.")
            else:
                # 1️⃣ Create User
                user = User.objects.create_user(
                    username=username,
                    email=email,
                    password=password1
                )

                # 2️⃣ Create StudentExtra
                student = StudentExtra(
                    user=user,
                    mobile_number=mobile_number,
                    roll_number=roll_number,
                    course=course,
                    department=department,
                    year=year
                )
                student.save()   # will trigger your custom save() → barcode created

                messages.success(request, "✅ Student added successfully!")
                return redirect('view_students')
        else:
            messages.error(request, "❌ Please correct the errors below.")
    else:
        form = StudentSignupForm()

    return render(
        request,
        'library/student_signup.html',
        {
            'form': form,
            'department_map_json': json.dumps(department_map, cls=DjangoJSONEncoder),
        }
    )


@staff_member_required
def load_departments(request):
    course_id = request.GET.get('course_id')
    departments = Department.objects.filter(course_id=course_id).values('id', 'name')
    return JsonResponse(list(departments), safe=False)
 


def student_login(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        
        if user is not None and not user.is_staff:  # Ensure it's a student
            login(request, user)

            # Redirect to 'next' if it exists (e.g., login from book detail)
            next_url = request.POST.get('next')
            if next_url:
                return redirect(next_url)

            # Default redirect to home
            return redirect('home')

        else:
            return render(request, 'library/student_login.html', {
                'error': 'Invalid student credentials'
            })

    # For GET request, capture any ?next=... parameter from URL
    next_url = request.GET.get('next', '')
    return render(request, 'library/student_login.html', {'next': next_url})

# third
def forgot_password(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')

        if not username or not email:
            messages.error(request, "Please enter both username and email.")
            return render(request, 'library/forgot_password.html')

        try:
            user = User.objects.get(username=username)
            student_extra = StudentExtra.objects.get(user=user)

            if user.email != email:
                messages.error(request, "The email doesn't match the registered email.")
                return render(request, 'library/forgot_password.html')

            # Email matches — proceed to send reset link
            uid = urlsafe_base64_encode(force_bytes(user.pk))
            token = default_token_generator.make_token(user)
            reset_url = request.build_absolute_uri(f"/reset-password/{uid}/{token}/")

            send_mail(
                subject="🔐 Password Reset | Library",
                message=f"Hello {user.username},\n\nClick the link below to reset your password:\n{reset_url}\n\nIf you didn’t request this, please ignore this email.",
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[email],
                fail_silently=False,
            )

            messages.success(request, "Password reset link has been sent to your registered email.")
        except User.DoesNotExist:
            messages.error(request, "No user found with that username.")
        except StudentExtra.DoesNotExist:
            messages.error(request, "No student profile found for this user.")

    return render(request, 'library/forgot_password.html')

def reset_password(request, uidb64, token):
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
    except (User.DoesNotExist, ValueError):
        user = None

    validlink = True  

    if request.method == 'POST':
        username = request.POST.get('username')
        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')

        if password1 != password2:
            messages.error(request, "❌ Passwords do not match.")
        else:
            try:
                user = User.objects.get(username=username)
                user.password = make_password(password1)
                user.save()
                messages.success(request, "✅ Password successfully updated.")
                return redirect('student_login')  # redirect after successful reset
            except User.DoesNotExist:
                messages.error(request, "❌ User not found.")

    return render(request, 'library/reset_password.html', {'validlink': validlink})
    

def student_dashboard(request):
    
    student = request.user
    
   
    unset_old_newly_published_books()
    send_expiry_reminders()
    # Optional: use Django messages to show thank you
    if request.GET.get('donated') == '1':
        messages.success(request, '🎉 Thank you for donating the book!')

    if request.method == "POST":
        form = StudentQueryForm(request.POST)
        if form.is_valid():
            query = form.save(commit=False)
            query.student = request.user
            query.save()
            messages.success(request, "Your query has been submitted.")
            return redirect('student_dashboard')
    else:
        form = StudentQueryForm()

     # Add this
    today = timezone.now().date()
    one_month_ago = today - timedelta(days=30)

    bestsellers = Book.objects.filter(bestseller=True)
    newly_published = Book.objects.filter(
        newly_published=True,
        published_date__isnull=False,
        published_date__gte=one_month_ago
    )

    notifications = Notification.objects.filter(student=request.user, seen=False)
    donated_books = Book.objects.filter(donated_by=student)

    # Add this:
    replied_queries = StudentQuery.objects.filter(
        student=student,
        reply__isnull=False
    ).order_by('-submitted_at')

    return render(request, 'library/student_dashboard.html', {
        'form': form,
        'replied_queries': replied_queries,
        'bestsellers': bestsellers,
        'newly_published': newly_published,
        'donated_books': donated_books,
        'student': student,
        # 'user': user,
    })


@staff_member_required
def donate_book(request):
    if request.method == "POST":
        form = BookForm(request.POST, request.FILES)
        if form.is_valid():
            book = form.save(commit=False)
            book.donated_by = request.user
            book.is_donated = True
            book.save()
            messages.success(request, "🎉 Thank you! Your donated book has been added successfully.")
            return redirect('view_donated_book')
    else:
        form = BookForm()

    return render(request, 'library/donate_book.html', {'form': form})

@staff_member_required
def edit_donate_book(request, book_id):
    book = get_object_or_404(Book, id=book_id, is_donated=True)

    if request.method == "POST":
        form = BookForm(request.POST, request.FILES, instance=book)
        if form.is_valid():
            form.save()
            messages.success(request, "✏️ Donated book has been updated successfully.")
            return redirect('view_donated_books')
        else:
            # Debug: show why the form failed
            print("❌ Form errors:", form.errors)
            messages.error(request, "Please correct the errors below.")
    else:
        form = BookForm(instance=book)

    return render(
        request,
        "library/donate_book.html",
        {"form": form, "edit": True, "book": book},
    )


@login_required
def student_ordered_books(request):
    student_issues = IssuedBook.objects.filter(student=request.user).select_related('book')
    
    ordered_books = []
    for issue in student_issues:
        # Determine waiting position
        waiting_position = None
        if not issue.returned:
            queue = IssuedBook.objects.filter(book=issue.book, returned=False).order_by('issued_date')
            ids = list(queue.values_list('student__id', flat=True))
            if request.user.id in ids:
                waiting_position = ids.index(request.user.id) + 1

        ordered_books.append({
            'book': issue.book,
            'issued_date': issue.issued_date,
            'expiry_date': issue.expiry_date,
            'returned': issue.returned,
            'fine': issue.fine,
            'waiting_position': waiting_position
        })
        

    return render(request, 'library/student_ordered_books.html')

def student_logout(request):
    logout(request)
    return redirect('student_login')



@staff_member_required
def view_books(request):
    books = Book.objects.select_related('genre').all() 
    # books = Book.objects.all()
    query = request.GET.get('q', '').strip()
    export = request.GET.get("export", "")

    all_books = Book.objects.all()
    parsed_date = parse_date(query)

    if query:
        all_books = all_books.filter(
            Q(name__icontains=query) |
            Q(author__icontains=query) |
            Q(language__icontains=query) |
            Q(publisher__icontains=query) |
            Q(quantity__icontains=query) |
            (Q(published_date=parsed_date) if parsed_date else Q()) |
            Q(genre__name__icontains=query)
        )

    # ✅ Export Excel if requested
    if export == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Books"

        # Header row
        ws.append([
            "Title", "Author", "ISBN", "Genre", "Language", 
            "Publisher", "Quantity", "Description", "Published","Donated"
        ])

        for b in all_books:
            ws.append([
                b.name,
                b.author,
                b.isbn,
                b.genre.name if b.genre else "",
                b.language,
                b.publisher,
                b.quantity,
                b.description,
                b.published_date.strftime("%Y-%m-%d") if b.published_date else "",
                "Yes" if b.is_donated else "No"
            ])

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = 'attachment; filename=books.xlsx'
        wb.save(response)
        return response

    # ✅ Regular pagination logic
    regular_books = all_books.filter(is_donated=False).order_by('-id')
    donated_books = all_books.filter(is_donated=True).order_by('-id')

    paginator_regular = Paginator(regular_books, 10)
    paginator_donated = Paginator(donated_books, 10)

    page_number_regular = request.GET.get('regular_page')
    page_number_donated = request.GET.get('donated_page')

    books_regular = paginator_regular.get_page(page_number_regular)
    books_donated = paginator_donated.get_page(page_number_donated)

    context = {
        'books_regular': books_regular,
        'books_donated': books_donated,
        'search_query': query,
        "books": books
    }
    return render(request, 'library/view_books.html', context)



@staff_member_required
def view_donated_books(request):
    query = request.GET.get("q", "")
    donated_books = Book.objects.filter(is_donated=True).order_by("-id")

    if query:
        donated_books = donated_books.filter(
            Q(name__icontains=query) |
            Q(author__icontains=query) |
            Q(isbn__icontains=query) |
            Q(description__icontains=query) |
            Q(donated_by__username__icontains=query) |
            Q(language__icontains=query) |
            Q(genre__name__icontains=query)
        )

    # books_regular = Book.objects.filter(is_donated=False).order_by("-id")
    paginator = Paginator(donated_books, 10)
    page = request.GET.get("page")
    books_page = paginator.get_page(page)
    
    print("✅ View reached, rendering template...") 

    return render(
        request,
        "library/view_donated_books.html",
        {"books_donated": books_page, "search_query": query,},
    )


@staff_member_required
def export_donated_books(request):
    query = request.GET.get("q", "")
    donated_books = Book.objects.filter(is_donated=True).order_by("-id")

    if query:
        donated_books = donated_books.filter(
            Q(name__icontains=query) |
            Q(author__icontains=query) |
            Q(isbn__icontains=query) |
            Q(description__icontains=query) |
            Q(donated_by__username__icontains=query) |
            Q(language__icontains=query) |
            Q(genre__name__icontains=query)
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Donated Books"

    # Header
    ws.append(["Book Name", "Author", "Genre", "ISBN", "Description", "Donated By"])

    # Rows
    for book in donated_books:
        ws.append([
            book.name,
            book.author,
            book.genre.name if book.genre else "",
            book.isbn,
            book.description if book.description else "",
            book.donated_by.username if book.donated_by else "",
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="donated_books.xlsx"'
    wb.save(response)
    return response

@staff_member_required
def edit_book(request, pk):
    book = get_object_or_404(Book, pk=pk)
    old_quantity = book.quantity  # Store old quantity

    if request.method == 'POST':
        form = BookForm(request.POST, request.FILES, instance=book)
        if form.is_valid():
            updated_book = form.save()  # ✅ Let model save() handle QR code
            new_quantity = updated_book.quantity

            # 🔁 Promote Reserved → Borrowed if quantity increased
            if new_quantity > old_quantity:
                borrowed_count = IssuedBook.objects.filter(book=book, status='Borrowed', returned=False).count()
                available_slots = new_quantity - borrowed_count

                if available_slots > 0:
                    reserved_records = IssuedBook.objects.filter(
                        book=book,
                        status='Reserved',
                        returned=False
                    ).order_by('issued_date')[:available_slots]

                    for rec in reserved_records:
                        rec.status = 'Borrowed'
                        rec.issued_date = timezone.now()
                        rec.expiry_date = timezone.now() + timedelta(days=7)
                        rec.save()

                    messages.success(request, f"{len(reserved_records)} reserved users upgraded to Borrowed.")

            messages.success(request, "✅ Book updated successfully.")
            return redirect('view_books')

        else:
            messages.error(request, "❌ Invalid form submission. Please check your fields.")
    else:
        form = BookForm(instance=book)

    return render(request, 'library/edit_book.html', {'form': form, 'book': book})


@staff_member_required
def delete_book(request, pk):
    book = get_object_or_404(Book, pk=pk)
    book.delete()
    return redirect('view_books')

@staff_member_required
def issue_book(request):
    return render(request, 'library/issue_book.html')

@staff_member_required
def view_issued_books(request):
    return render(request, 'library/view_issued_books.html')


@staff_member_required
def export_students_excel(request):
    query = request.GET.get('q', '').strip()
    selected_department = request.GET.get('department', '').strip()

    # ✅ Base queryset
    students_qs = User.objects.filter(studentextra__isnull=False) \
        .select_related('studentextra__course', 'studentextra__department') \
        .order_by('username')

    # ✅ Filter by department
    if selected_department:
        if selected_department.isdigit():
            students_qs = students_qs.filter(studentextra__department_id=int(selected_department))
        else:
            students_qs = students_qs.filter(studentextra__department__name__iexact=selected_department)

    # ✅ Search filter
    if query:
        students_qs = students_qs.filter(
            Q(username__icontains=query) |
            Q(email__icontains=query) |
            Q(studentextra__mobile_number__icontains=query) |
            Q(studentextra__roll_number__icontains=query) |
            Q(studentextra__department__name__icontains=query) |
            Q(studentextra__course__name__icontains=query) |
            Q(studentextra__year__icontains=query)
        )

    # ✅ Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"

    # Headers
    ws.append(["Username", "Email", "Roll No", "Course", "Department", "Year", "Mobile"])

    # Rows
    for s in students_qs:
        ws.append([
            s.username,
            s.email,
            getattr(s.studentextra, "roll_number", ""),
            getattr(s.studentextra.course, "name", "") if s.studentextra.course else "",
            getattr(s.studentextra.department, "name", "") if s.studentextra.department else "",
            getattr(s.studentextra, "year", ""),
            getattr(s.studentextra, "mobile_number", ""),
        ])

    # Response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="students.xlsx"'
    wb.save(response)
    return response

@staff_member_required
def view_students(request):
    query = request.GET.get('q', '').strip()
    selected_department = request.GET.get('department', '').strip()

    # ✅ Ensure all StudentExtra rows have a department
    default_department = Department.objects.first()
    if default_department:
        StudentExtra.objects.filter(department__isnull=True).update(department=default_department)

    # ✅ Base queryset: only users with studentextra
    students_qs = User.objects.filter(studentextra__isnull=False) \
        .select_related('studentextra__course', 'studentextra__department') \
        .order_by('username')

    # ✅ Filter by department
    if selected_department:
        if selected_department.isdigit():
            students_qs = students_qs.filter(studentextra__department_id=int(selected_department))
        else:
            students_qs = students_qs.filter(studentextra__department__name__iexact=selected_department)

    # ✅ Search filter
    if query:
        students_qs = students_qs.filter(
            Q(username__icontains=query) |
            Q(email__icontains=query) |
            Q(studentextra__mobile_number__icontains=query) |
            Q(studentextra__roll_number__icontains=query) |
            Q(studentextra__department__name__icontains=query) |
            Q(studentextra__course__name__icontains=query) |
            Q(studentextra__year__icontains=query)
        )

    # ✅ Order students
    # students_qs = students_qs.order_by('username')
    # ✅ Order students: latest first
    students_qs = students_qs.order_by('-date_joined')


    # ✅ Get all queries for these students
    student_ids = students_qs.values_list('id', flat=True)
    all_queries = (
        StudentQuery.objects
        .select_related('student')
        .filter(student_id__in=student_ids)
    )

    # ✅ Group queries by student
    queries_by_student = {}
    for q in all_queries:
        q.reply_form = ReplyForm(instance=q) if not q.reply else None
        queries_by_student.setdefault(q.student_id, []).append(q)

    for student in students_qs:
        student.queries = queries_by_student.get(student.id, [])

    # ✅ Pagination
    paginator = Paginator(students_qs, 10)
    students_page = paginator.get_page(request.GET.get('page'))

    context = {
        'students': students_page,
        'query': query,
        'departments': Department.objects.all(),
        'selected_department': selected_department,
    }
    return render(request, 'library/view_students.html', context)

@staff_member_required
def delete_issue(request, pk):
    issue = get_object_or_404(IssuedBook, pk=pk)
    issue.delete()
    return redirect('view_students')

@staff_member_required
def delete_student(request, user_id):
    user = get_object_or_404(User, id=user_id, is_staff=False)
    user.delete()
    messages.success(request, "Student deleted successfully.")
    return redirect('view_students')


@login_required
def student_issued_books(request):
    student = request.user
    issued_books = IssuedBook.objects.filter(student=student).order_by('-issued_date')
    # returned_books = IssuedBook.objects.filter(student=student, returned=True)
    canceled_books = IssuedBook.objects.filter(student=student, canceled=True)

    borrowed_books = IssuedBook.objects.filter(student=student, status='Borrowed', returned=False)
    reserved_books = IssuedBook.objects.filter(student=student, status='Reserved', returned=False)
    waiting_books = IssuedBook.objects.filter(student=student, status='Waiting', returned=False)
    returned_books = IssuedBook.objects.filter(student=student, returned=True)

    return render(request, 'library/student_issued_books.html', {
        'borrowed_books': borrowed_books,
        'reserved_books': reserved_books,
        'waiting_books': waiting_books,
        'returned_books': returned_books,
        'student': student,
        'issued_books': issued_books,
        'canceled_books': canceled_books,
      

    })


# For admin
@staff_member_required
def admin_canceled_books(request):
    canceled_books = IssuedBook.objects.filter(is_canceled=True).order_by('-issued_date')
    return render(request, 'library/view_issued_books.html', {'canceled_books': canceled_books})





@staff_member_required
def edit_student(request, student_id):
    user = get_object_or_404(User, id=student_id)
    student_extra, created = StudentExtra.objects.get_or_create(user=user)

    if request.method == 'POST':
        user_form = EditStudentForm(request.POST, instance=user)
        extra_form = EditStudentExtraForm(request.POST, instance=student_extra)

        if user_form.is_valid() and extra_form.is_valid():
            user_form.save()
            extra_form.save()
            messages.success(request, "Student details updated successfully!")
            return redirect('view_students')
    else:
        user_form = EditStudentForm(instance=user)
        extra_form = EditStudentExtraForm(instance=student_extra)

    return render(request, 'library/edit_student.html', {
        'user_form': user_form,
        'extra_form': extra_form,
        'user_id': user.id
    })



@staff_member_required
def send_reply(request, query_id):  # <-- Fix here
    try:
        query = StudentQuery.objects.get(pk=query_id)
    except StudentQuery.DoesNotExist:
        messages.error(request, "Query not found.")
        return redirect('view_students')

    if request.method == 'POST':
        reply = request.POST.get('reply')
        if reply:
            query.reply = reply
            query.save()
            messages.success(request, "Reply sent successfully.")
        else:
            messages.warning(request, "Reply message is empty.")

    return redirect('view_students')



@login_required
def cancel_reservation(request, issued_book_id):
    issued_book = get_object_or_404(IssuedBook, id=issued_book_id, student=request.user)

    if issued_book.status in ['Reserved', 'Waiting']:
        issued_book.delete()
        messages.success(request, f"{issued_book.book.name} reservation/waiting canceled.")
        # call your promotion logic
        promote_users(issued_book.book)
    else:
        messages.warning(request, "You can only cancel reserved or waiting books.")

    print("Issued Book ID:", issued_book_id)
    return redirect('student_issued_books')  # adjust the name as per your project




def add_book(request):
   
    if request.method == 'POST':
        form = BookForm(request.POST, request.FILES)
        if form.is_valid():
            # form.save()
            book = form.save()
            
            messages.success(request, "✍️ Book added successfully!")
            return redirect('view_books') 

    else:
        form = BookForm()
    return render(request, 'library/add_book.html', {'form': form})


@staff_member_required
def issue_book(request):
    book_id = request.GET.get('book_id')
    username = request.GET.get('username')

    if request.method == 'POST':
        student_username = request.POST.get('username')
        book_id_post = request.POST.get('book_id')  # from scanner / hidden field
        book_title = request.POST.get('book_title')  # from manual entry

        # optional dates
        issued_date_str = request.POST.get('issued_date')
        returned_date_str = request.POST.get('returned_date')

        issued_date = timezone.now().date()
        if issued_date_str:
            try:
                issued_date = timezone.datetime.strptime(issued_date_str, '%Y-%m-%d').date()
            except ValueError:
                messages.error(request, "Invalid issued date format. Using today as default.")

        returned_date = timezone.now().date()
        if returned_date_str:
            try:
                returned_date = timezone.datetime.strptime(returned_date_str, '%Y-%m-%d').date()
            except ValueError:
                messages.error(request, "Invalid return date format. Using today as default.")

        # Validate student
        if not student_username:
            messages.error(request, "Missing student username.")
            return redirect('issue_book')

        try:
            student = User.objects.get(username=student_username)
        except User.DoesNotExist:
            messages.error(request, f"Student '{student_username}' not found.")
            return redirect('issue_book')

        # Validate book
        if book_id_post:
            try:
                book = Book.objects.get(id=book_id_post)
            except Book.DoesNotExist:
                messages.error(request, "Book not found.")
                return redirect('issue_book')
        else:
            # Find first available copy of the book
            available_books = Book.objects.filter(
                name__iexact=book_title
            ).exclude(
                id__in=IssuedBook.objects.filter(returned=False).values_list('book_id', flat=True)
            ).order_by('copy_number')

            if not available_books.exists():
                messages.warning(request, f"All copies of '{book_title}' are already issued or reserved.")
                return redirect('issue_book')

            book = available_books.first()

        # Check if student already has this copy
        if IssuedBook.objects.filter(book=book, student=student, returned=False).exists():
            messages.warning(request, f"{book.name} (Copy {book.copy_number}) is already issued to {student.username}.")
            return redirect('student_issued_books' if not request.user.is_superuser else 'view_issued_books')

        # Library settings
        settings = LibrarySetting.get_settings()
        issued_qs = IssuedBook.objects.filter(book__name=book.name, returned=False).order_by('issued_date')
        total_copies = Book.objects.filter(name=book.name).count()
        borrowed_count = issued_qs.filter(status='Borrowed').count()
        reserved_count = issued_qs.filter(status='Reserved').count()
        available_copies = total_copies - borrowed_count

        # Expiry days
        custom_days = request.POST.get('custom_expiry_days')
        custom_days = int(custom_days) if custom_days else settings.expiry_days

        # Issue / Reserve / Waiting logic
        if available_copies > 0:
            # Issue
            IssuedBook.objects.create(
                student=student,
                book=book,
                status='Borrowed',
                issued_date=issued_date,
                returned_date=returned_date,
                expiry_date=issued_date + timedelta(days=custom_days),
                custom_expiry_days=custom_days,
            )
            messages.success(request, f"{book.name} (Copy {book.copy_number}) successfully issued to {student.username}.")

        elif reserved_count < total_copies:
            # Reserve
            IssuedBook.objects.create(
                student=student,
                book=book,
                status='Reserved',
                issued_date=issued_date,
                returned_date=returned_date,
                expiry_date=issued_date + timedelta(days=settings.expiry_days),
                custom_expiry_days=custom_days,
            )
            messages.info(request, f"{book.name} (Copy {book.copy_number}) reserved for {student.username}.")

        else:
            # Waiting
            waiting_qs = issued_qs.filter(status='Waiting')
            position = waiting_qs.count()
            wait_blocks = (position // total_copies) + 1
            wait_days = wait_blocks * settings.expiry_days

            IssuedBook.objects.create(
                student=student,
                book=book,
                status='Waiting',
                issued_date=issued_date,
                returned_date=returned_date,
                expiry_date=issued_date + timedelta(days=wait_days),
                custom_expiry_days=wait_days,
            )
            messages.warning(
                request,
                f"{book.name} (Copy {book.copy_number}) is fully booked. "
                f"{student.username} is added to the waitlist (position {position + 1}). "
                f"Estimated wait: {wait_days} days."
            )

        return redirect('student_issued_books' if not request.user.is_superuser else 'view_issued_books')

    # GET method → preload student & book if scanned
    initial_data = {}
    if book_id and username:
        book = get_object_or_404(Book, id=book_id)
        student = get_object_or_404(User, username=username)
        initial_data = {
            'book_id': book.id,
            'book_title': book.name,
            'username': student.username
        }

    return render(request, 'library/issue_book.html', {
        'is_student': bool(book_id and username),
        'initial': initial_data
    })


def fetch_student_by_barcode(request):
    barcode_id = request.GET.get('code')

    if not barcode_id:
        return JsonResponse({'error': 'No barcode_id provided'}, status=400)

    student = StudentExtra.objects.filter(barcode_id=barcode_id).first()

    if student:
        return JsonResponse({
            
            'username': student.user.username,
            'roll_number': student.roll_number,
            'mobile': student.mobile_number,
            'email': student.user.email,
            
        })

    return JsonResponse({'error': 'Student not found'}, status=404)


def scan_lookup(request):
    code = request.GET.get('code')
    if not code:
        return JsonResponse({'error': 'No code provided'}, status=400)

    # Check Student Barcode
    try:
        student = StudentExtra.objects.select_related('user').get(barcode_id=code)
        return JsonResponse({
            'username': student.user.username,
            'roll_number': student.roll_number,
            'mobile_number': student.mobile_number,
            'email': student.user.email
        })
    except StudentExtra.DoesNotExist:
        pass
    
    if code.startswith("BOOK|"):
        try:
            parts = code.split('|')
            if len(parts) >= 3:
                title = parts[1].strip()
                isbn = parts[2].strip()

                book = Book.objects.filter(title__iexact=title, isbn__iexact=isbn).first()
                if book:
                    return JsonResponse({
                        'type': 'book',
                        'title': book.name,
                        'isbn': book.isbn,
                        'author': book.author,
                        'genre': book.genre,
                        'language': book.language,
                    })
                else:
                    return JsonResponse({'error': 'Book not found'}, status=404)
            else:
                return JsonResponse({'error': 'Invalid book QR format'}, status=400)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'No match found'}, status=404)


def api_scan_lookup(request):
    code = request.GET.get('code')
    if not code:
        return JsonResponse({'error': 'No code provided.'})

    try:
        # Assuming barcode holds student's roll_number or username, adjust as per your model
        user = User.objects.get(studentextra__roll_number=code)
        return JsonResponse({
            'username': user.get_full_name() or user.username,
            'roll_number': user.studentextra.roll_number
        })
    except User.DoesNotExist:
        return JsonResponse({'error': 'Student not found.'})
    
from django.utils.timezone import now
from django.http import JsonResponse
from .models import IssuedBook

def api_check_issue(request):
    roll = request.GET.get('roll')
    isbn = request.GET.get('isbn')

    if not roll or not isbn:
        return JsonResponse({'error': 'Missing parameters.'})

    try:
        issued_book = IssuedBook.objects.select_related('student', 'book').get(
            student__studentextra__roll_number=roll,
            book__isbn=isbn,
            returned=False
        )

        # ✅ Use the same methods everywhere
        fine = issued_book.calculate_fine()
        display_fine = issued_book.get_display_fine()

        return JsonResponse({
            'fine': fine,  # numeric, for logic
            'display_fine': display_fine,  # formatted string (₹80)
            'username': issued_book.student.get_full_name() or issued_book.student.username,
            'issued_book_id': issued_book.id,
            "book_name": issued_book.book.name,
        })

    except IssuedBook.DoesNotExist:
        return JsonResponse({'error': 'No issued record found for this student and book.'})


# def api_check_issue(request):
#     roll = request.GET.get('roll')
#     isbn = request.GET.get('isbn')
   

#     if not roll or not isbn:
#         return JsonResponse({'error': 'Missing parameters.'})

#     try:
#         issued_book = IssuedBook.objects.select_related('student', 'book').get(
#             student__studentextra__roll_number=roll,
#             book__isbn=isbn,
#             returned=False
#         )

#         # Calculate fine dynamically if not returned
#         fine = 0
#         if issued_book.expiry_date and now().date() > issued_book.expiry_date:
#             days_late = (now().date() - issued_book.expiry_date).days
#             fine = days_late * issued_book.fine_per_day

#         return JsonResponse({
#             'fine': fine,
#             'username': issued_book.student.get_full_name() or issued_book.student.username,
#             'issued_book_id': issued_book.id,
#         })

#     except IssuedBook.DoesNotExist:
#         return JsonResponse({'error': 'No issued record found for this student and book.'})


from django.views.decorators.http import require_http_methods
from django.db import transaction

@require_http_methods(["GET", "POST"])
def return_book(request):
    if request.method == "POST":
        roll = request.POST.get('roll_number')
        isbn = request.POST.get('isbn')
        payment_method = request.POST.get('payment_method', '')

        if not (roll and isbn):
            messages.error(request, "Please provide student roll number and book ISBN.")
            return redirect('return_book')

        try:
            with transaction.atomic():
                issued_book = IssuedBook.objects.select_for_update().get(
                    student__studentextra__roll_number=roll,
                    book__isbn=isbn,
                    returned=False
                )

                calculated_fine = issued_book.calculate_fine()

                if payment_method == "CanceledByAdmin":
                    # ✅ Fine canceled by admin due to holiday
                    issued_book.final_fine = calculated_fine
                    issued_book.fine_cancelled = True
                    issued_book.fine_cancel_reason = "College holiday considered"
                    issued_book.payment_method = None
                else:
                    # ✅ Normal fine flow
                    if calculated_fine > 0 and not payment_method:
                        messages.error(request, "Fine pending. Please select a payment method before returning the book.")
                        return redirect('return_book')

                    issued_book.final_fine = calculated_fine
                    issued_book.payment_method = payment_method if calculated_fine > 0 else None
                    issued_book.fine_cancelled = False
                    issued_book.fine_cancel_reason = None

                # ✅ Mark as returned
                issued_book.returned = True
                issued_book.returned_date = now().date()
                issued_book.status = 'Returned'
                issued_book.save()

                # ✅ Update book quantity
                book = issued_book.book
                book.quantity += 1
                book.save()

                # ✅ Success message
                msg = f"✅ Book '{book.name}' returned successfully!"
                if issued_book.fine_cancelled:
                    msg += " (Fine waived — College holiday considered)"
                elif issued_book.final_fine > 0:
                    msg += f" Fine collected: ₹{issued_book.final_fine}"
                messages.success(request, msg)

                return redirect('view_issued_books')

        except IssuedBook.DoesNotExist:
            messages.error(request, "No active issued book record found for this student and book.")
            return redirect('return_book')

    return render(request, 'library/return_book.html')

@csrf_exempt
@require_POST
def issue_book_api(request):
    username     = request.POST.get("username")
    roll_number  = request.POST.get("roll_number")
    email        = request.POST.get("email")
    mobile_number  = request.POST.get("mobile_number")
    isbn         = request.POST.get("isbn")
    title        = request.POST.get("name")
    genre        = request.POST.get("genre")
    language     = request.POST.get("language")

    if not (username and roll_number and isbn and title and genre and language and email and mobile_number):
        return JsonResponse({"ok": False, "error": "Incomplete form data."}, status=400)

    try:
        student = StudentExtra.objects.get(user__username=username, roll_number=roll_number,    user__email=email, mobile_number=mobile_number)
    except StudentExtra.DoesNotExist:
        return JsonResponse({"ok": False, "error": "Student not found."}, status=404)

    try:
        book = Book.objects.get(isbn=isbn)
    except Book.DoesNotExist:
        return JsonResponse({"ok": False, "error": "Book not found."}, status=404)

    issued_qs = IssuedBook.objects.filter(book=book, returned=False)
    borrowed_count = issued_qs.filter(status='Borrowed').count()
    reserved_count = issued_qs.filter(status='Reserved').count()
    available = book.quantity - borrowed_count

    expiry_days = int(request.POST.get("custom_expiry_days", 7))

    if available > 0:
        IssuedBook.objects.create(
            student=student.user,
            book=book,
            status='Borrowed',
            issued_date=timezone.now(),
            expiry_date=timezone.now() + timedelta(days=expiry_days),
        )
        return JsonResponse({"ok": True, "message": f"Book '{book.name}' issued to {student.user.username}."})

    elif reserved_count < book.quantity:
        IssuedBook.objects.create(
            student=student.user,
            book=book,
            status='Reserved',
            issued_date=timezone.now(),
            expiry_date=timezone.now() + timedelta(days=expiry_days),
        )
        return JsonResponse({"ok": True, "message": f"Book '{book.name}' reserved for {student.user.username}."})

    else:
        waiting_qs = issued_qs.filter(status='Waiting')
        position = waiting_qs.count()
        wait_days = ((position // book.quantity) + 1) * 7

        IssuedBook.objects.create(
            student=student.user,
            book=book,
            status='Waiting',
            issued_date=timezone.now(),
            expiry_date=timezone.now() + timedelta(days=wait_days),
        )
        return JsonResponse({
            "ok": True,
            "message": f"Book fully booked. {student.user.username} is added to waitlist (pos {position+1}, wait {wait_days} days)."
        })

def search_books_ajax(request):
    term = request.GET.get('term', '')
    results = []

    if term:
        books = Book.objects.filter(name__icontains=term)[:10]
        for book in books:
            results.append({
                'label': f"{book.name} ({'Available' if book.quantity > 0 else 'Not Available'})",
                'value': book.name
            })

    return JsonResponse(results, safe=False)


@staff_member_required
@csrf_exempt   # since you're using fetch
def mark_payment_received(request):
    if request.method == "POST":
        data = json.loads(request.body)
        record_id = data.get("record_id")
        method = data.get("method")
        amount = data.get("amount")
        action = data.get("action")

        issued_book = get_object_or_404(IssuedBook, id=record_id)

        if action == "missing":
            penalty = issued_book.book.price + issued_book.final_fine
            issued_book.status = "Missing"
            issued_book.missing = True
            issued_book.final_fine = penalty
            issued_book.payment_method = method
            issued_book.save()

            # reduce stock
            if issued_book.book.quantity > 0:
                issued_book.book.quantity -= 1
                issued_book.book.save()

        elif action == "renew":
            issued_book.payment_method = method
            issued_book.final_fine = amount
            issued_book.save()

        return JsonResponse({"status": "success"})

    return JsonResponse({"status": "error", "msg": "Invalid request"}, status=400)

from django.db.models.functions import Coalesce
from django.db.models import Q
@staff_member_required
def view_issued_books(request):
    issued_books = IssuedBook.objects.select_related('student', 'book')

    today = timezone.now().date()
    expiry_messages = []
    for issued in issued_books:
        if issued.expiry_date and issued.expiry_date < today and not issued.returned:
            expiry_messages.append(
                f"⚠️ Book '{issued.book.name}' issued to {issued.student.username} expired on {issued.expiry_date}"
            )

    # --- Filters ---
    q = request.GET.get('q', '').strip()
    from_date_str = request.GET.get('from_date', '')
    to_date_str = request.GET.get('to_date', '')
    tab = request.GET.get('tab', 'borrowed')
    entries_per_page = int(request.GET.get('entries', 10) or 10)

    parsed_date = parse_date(q)

    def apply_search(qs):
        if q:
            qs = qs.filter(
                Q(book__name__icontains=q) |
                Q(book__author__icontains=q) |
                Q(book__genre__name__icontains=q) |
                Q(book__isbn__icontains=q) |
                Q(student__username__icontains=q) |
                Q(student__email__icontains=q) |
                Q(canceled_by__icontains=q) |
                (Q(issued_date=parsed_date) if parsed_date else Q()) |
                (Q(expiry_date=parsed_date) if parsed_date else Q()) |
                (Q(returned_date=parsed_date) if parsed_date else Q())
            )
        if from_date_str:
            fd = parse_date(from_date_str)
            if fd:
                qs = qs.filter(issued_date__gte=fd)
        if to_date_str:
            td = parse_date(to_date_str)
            if td:
                qs = qs.filter(issued_date__lte=td)
        return qs

   
    def filter_issued_books(status):
        qs = IssuedBook.objects.select_related('student', 'book').filter(
            status=status,
            canceled=False
        )

        # ✅ For Borrowed tab → only keep active borrowings
        if status == "Borrowed":
            qs = qs.filter(returned=False)

        return apply_search(qs).order_by('-issued_date')

    borrowed_qs = filter_issued_books('Borrowed')
    reserved_qs = filter_issued_books('Reserved')
    waiting_qs  = filter_issued_books('Waiting')
    missing_qs  = filter_issued_books('Missing')
    
    canceled_qs = apply_search(
        IssuedBook.objects.select_related('student','book').filter(canceled=True)
    ).order_by('-issued_date')

    returned_qs = apply_search(
        IssuedBook.objects.select_related('student','book').filter(returned=True, canceled=False)
    ).order_by('-returned_date')

    # --- Fines Queryset ---
    fines_qs = apply_search(
        IssuedBook.objects.select_related('student', 'book').filter(
            Q(final_fine__gt=0) |                 # fines exist
            Q(payment_method__isnull=False) |     # paid
            Q(fine_cancel_reason__isnull=False)   # canceled
        )
    )
    # Latest first: by returned_date if exists, else issued_date
    fines_qs = fines_qs.order_by(
        Coalesce('returned_date', 'issued_date').desc()
    )
    # fines_qs = apply_search(
    #     IssuedBook.objects.select_related('student', 'book').filter(
    #         Q(final_fine__gt=0) |                  # actual fines
    #         Q(payment_method__isnull=False) |      # paid fines
    #         Q(fine_cancel_reason__isnull=False)    # canceled fines (holiday etc.)
    #     )
    # ).order_by('-issued_date')


    
    # --- Choose tab ---
    data = {
        'borrowed': borrowed_qs,
        'reserved': reserved_qs,
        'waiting':  waiting_qs,
        'returned': returned_qs,
        'canceled': canceled_qs,
        'fines':    fines_qs,
        'missing':  missing_qs
    }.get(tab, borrowed_qs)

    # --- Pagination ---
    paginator = Paginator(data, entries_per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'tab_list': [
            ('borrowed', 'Borrowed'),
            ('reserved', 'Reserved'),
            ('waiting',  'Waiting'),
            ('returned', 'Returned'),
            ('canceled', 'Canceled'),
            ('fines',    'Fines'),
            ('missing',  'Missing'), 
        ],
        'active_tab': tab,
        'query': q,
        'from_date': from_date_str,
        'to_date': to_date_str,
        'page_obj': page_obj,
        'issued_books': issued_books,
        'expiry_messages': expiry_messages,
    }
    return render(request, 'library/view_issued_books.html', context)


@staff_member_required
def export_issued_books_view(request):
    """
    Rebuild the SAME filtered queryset used on the page and export it.
    """
    q = request.GET.get('q', '').strip()
    from_date_str = request.GET.get('from_date', '')
    to_date_str   = request.GET.get('to_date', '')
    tab = request.GET.get('tab', 'borrowed')

    parsed_date = parse_date(q)

    def apply_search(qs):
        if q:
            qs = qs.filter(
                Q(book__name__icontains=q) |
                Q(book__author__icontains=q) |
                Q(book__genre__name__icontains=q) |
                Q(book__isbn__icontains=q) |
                Q(student__username__icontains=q) |
                Q(student__email__icontains=q) |
                Q(canceled_by__icontains=q) |
                (Q(issued_date=parsed_date) if parsed_date else Q()) |
                (Q(expiry_date=parsed_date) if parsed_date else Q()) |
                (Q(returned_date=parsed_date) if parsed_date else Q())
            )
        if from_date_str:
            fd = parse_date(from_date_str)
            if fd:
                qs = qs.filter(issued_date__gte=fd)
        if to_date_str:
            td = parse_date(to_date_str)
            if td:
                qs = qs.filter(issued_date__lte=td)
        return qs

    def filter_issued_books(status):
        qs = IssuedBook.objects.select_related('student','book').filter(status=status, canceled=False)
        return apply_search(qs).order_by('-issued_date')

    borrowed_qs = filter_issued_books('Borrowed')
    reserved_qs = filter_issued_books('Reserved')
    waiting_qs  = filter_issued_books('Waiting')

    canceled_qs = apply_search(
        IssuedBook.objects.select_related('student','book').filter(canceled=True)
    ).order_by('-issued_date')

    returned_qs = apply_search(
        IssuedBook.objects.select_related('student','book').filter(returned=True, canceled=False)
    ).order_by('-returned_date')

    fines_qs = apply_search(
        IssuedBook.objects.select_related('student','book')
        .filter(canceled=False)
        .filter(Q(returned=True, final_fine__gt=0) | Q(returned=False, expiry_date__lt=date.today()))
    ).order_by('-issued_date')

    data = {
        'borrowed': borrowed_qs,
        'reserved': reserved_qs,
        'waiting':  waiting_qs,
        'returned': returned_qs,
        'canceled': canceled_qs,
        'fines':    fines_qs
    }.get(tab, borrowed_qs)

    return _export_issued_books_qs(data)


def _export_issued_books_qs(qs):
    """
    Helper that actually writes the Excel from a queryset.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Issued Books"

    # Header row
    ws.append([
        "Student", "Book", "Author", "Genre", "ISBN",
        "Issued Date", "Expiry Date", "Returned Date", "Status", "Final Fine"
    ])

    for b in qs:
        ws.append([
            (b.student.username if b.student_id else ""),
            (b.book.name       if b.book_id else ""),
            (b.book.author     if b.book_id else ""),
            (b.book.genre.name if b.book_id and getattr(b.book, "genre_id", None) else ""),
            (b.book.isbn       if b.book_id else ""),
            (b.issued_date.strftime("%Y-%m-%d")  if getattr(b, "issued_date", None)  else ""),
            (b.expiry_date.strftime("%Y-%m-%d")  if getattr(b, "expiry_date", None)  else ""),
            (b.returned_date.strftime("%Y-%m-%d") if getattr(b, "returned_date", None) else ""),
            (b.status or ""),
            (getattr(b, "final_fine", 0) or 0),
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="issued_books.xlsx"'
    wb.save(response)
    return response

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils.timezone import now
from datetime import timedelta
from .models import IssuedBook, Book, LibrarySetting

# @staff_member_required
# def admin_renew_book(request, issued_book_id):
#     issued_book = get_object_or_404(IssuedBook, id=issued_book_id)

#     # disallow if already renewed
#     if issued_book.renewed:
#         messages.error(request, "❌ You have already renewed this book once.")
#         return redirect('view_issued_books')

#     # disallow if reserved or waiting
#     if IssuedBook.objects.filter(book=issued_book.book, status__in=['Reserved', 'Waiting']).exists():
#         messages.error(request, "⚠️ Renewal not allowed, this book is reserved/waiting.")
#         return redirect('view_issued_books')

#     # calculate fine using model method
#     fine = issued_book.calculate_fine()

#     if request.method == "POST":
#         payment_method = request.POST.get("payment_method")
#         if fine > 0 and not payment_method:
#             messages.error(request, "⚠️ Please choose payment method to pay fine.")
#             return redirect('view_issued_books')

#         # reset fine after payment
#         issued_book.final_fine = 0
#         issued_book.payment_method = payment_method if fine > 0 else None
#         issued_book.expiry_date += timedelta(days=7)
#         issued_book.renewed = True
#         issued_book.save()

#         messages.success(request, f"✅ Book renewed successfully until {issued_book.expiry_date}")
#         return redirect('view_issued_books')

#     return redirect('view_issued_books')

from django.http import JsonResponse
from django.utils.timezone import now

@staff_member_required
def admin_renew_book(request, pk):
    issued_book = get_object_or_404(IssuedBook, id=pk)

    # disallow if already renewed
    if issued_book.renewed:
        return JsonResponse({"success": False, "message": "❌ You have already renewed this book once."})

    # disallow if reserved or waiting
    if IssuedBook.objects.filter(book=issued_book.book, status__in=['Reserved', 'Waiting']).exists():
        return JsonResponse({"success": False, "message": "⚠️ Renewal not allowed, this book is reserved/waiting."})

    fine = issued_book.calculate_fine()

    if request.method == "POST":
        payment_method = request.POST.get("payment_method")

        if payment_method == "CanceledByAdmin":
            # ✅ Fine is canceled, allow renewal
            issued_book.final_fine = 0
            issued_book.fine_cancelled = True
            issued_book.canceled_by = "admin"
            issued_book.fine_cancel_reason = "Fine waived during renewal (holiday/exception)"
            issued_book.payment_method = None
        else:
            # Normal payment case
            if fine > 0 and not payment_method:
                return JsonResponse({"success": False, "message": "⚠️ Please choose payment method to pay fine."})

            issued_book.final_fine = 0 if fine > 0 else fine
            issued_book.payment_method = payment_method if fine > 0 else None
            issued_book.fine_cancelled = False
            issued_book.fine_cancel_reason = None

        # ✅ Extend expiry & mark renewed
        issued_book.expiry_date += timedelta(days=7)
        issued_book.renewed = True
        issued_book.save()

        return JsonResponse({
            "success": True,
            "message": f"✅ Book renewed successfully until {issued_book.expiry_date}"
        })

    # for GET request (load modal data)
    return JsonResponse({
        "success": True,
        "fine": fine,
        "expiry_date": issued_book.expiry_date.strftime("%Y-%m-%d")
    })


# @staff_member_required
# def admin_renew_book(request, pk):
#     issued_book = get_object_or_404(IssuedBook, id=pk)

#     # disallow if already renewed
#     if issued_book.renewed:
#         return JsonResponse({"success": False, "message": "❌ You have already renewed this book once."})

#     # disallow if reserved or waiting
#     if IssuedBook.objects.filter(book=issued_book.book, status__in=['Reserved', 'Waiting']).exists():
#         return JsonResponse({"success": False, "message": "⚠️ Renewal not allowed, this book is reserved/waiting."})

#     fine = issued_book.calculate_fine()

#     if request.method == "POST":
#         payment_method = request.POST.get("payment_method")
#         if fine > 0 and not payment_method:
#             return JsonResponse({"success": False, "message": "⚠️ Please choose payment method to pay fine."})

#         # reset fine after payment
#         issued_book.final_fine = 0
#         issued_book.payment_method = payment_method if fine > 0 else None
#         issued_book.expiry_date += timedelta(days=7)
#         issued_book.renewed = True
#         issued_book.save()

#         return JsonResponse({
#             "success": True,
#             "message": f"✅ Book renewed successfully until {issued_book.expiry_date}"
#         })

#     # for GET request (load modal data)
#     return JsonResponse({
#         "success": True,
#         "fine": fine,
#         "expiry_date": issued_book.expiry_date.strftime("%Y-%m-%d")
#     })


# @staff_member_required
# def admin_renew_book(request, pk):
#     issued_book = get_object_or_404(IssuedBook, id=pk)

#     # already renewed
#     if issued_book.renewed:
#         messages.error(request, "❌ You have already renewed this book once.")
#         return redirect('view_issued_books')

#     # reserved/waiting not allowed
#     if IssuedBook.objects.filter(book=issued_book.book, status__in=['Reserved', 'Waiting']).exists():
#         messages.error(request, "⚠️ Renewal not allowed, this book is reserved/waiting.")
#         return redirect('view_issued_books')

#     fine = issued_book.calculate_fine()

#     if request.method == "POST" or fine == 0:
#         payment_method = request.POST.get("payment_method") if fine > 0 else None
#         if fine > 0 and not payment_method:
#             messages.error(request, "⚠️ Please choose payment method to pay fine.")
#             return redirect('view_issued_books')

#         issued_book.final_fine = 0
#         issued_book.payment_method = payment_method
#         issued_book.expiry_date += timedelta(days=7)
#         issued_book.renewed = True
#         issued_book.save()

#         messages.success(request, f"✅ Book renewed successfully until {issued_book.expiry_date}")
#         return redirect('view_issued_books')

#     return redirect('view_issued_books')


# @staff_member_required
# def admin_renew_book(request, pk):
#     issued_book = get_object_or_404(IssuedBook, id=pk)

#     # disallow if already renewed
#     if issued_book.renewed:
#         messages.error(request, "❌ You have already renewed this book once.")
#         return redirect('view_issued_books')

#     # disallow if reserved or waiting
#     if IssuedBook.objects.filter(book=issued_book.book, status__in=['Reserved', 'Waiting']).exists():
#         messages.error(request, "⚠️ Renewal not allowed, this book is reserved/waiting.")
#         return redirect('view_issued_books')

#     fine = issued_book.calculate_fine()

#     if request.method == "POST":
#         payment_method = request.POST.get("payment_method")
#         if fine > 0 and not payment_method:
#             messages.error(request, "⚠️ Please choose payment method to pay fine.")
#             return redirect('view_issued_books')

#         # reset fine after payment
#         issued_book.final_fine = 0
#         issued_book.payment_method = payment_method if fine > 0 else None
#         issued_book.expiry_date += timedelta(days=7)
#         issued_book.renewed = True
#         issued_book.save()

#         messages.success(request, f"✅ Book renewed successfully until {issued_book.expiry_date}")
#         return redirect('view_issued_books')

#     # Instead of rendering view_issued_books.html here, show a small confirm page
#     return render(request, "library/view_issued_books.html", {"issued_book": issued_book, "fine": fine})


@staff_member_required
def mark_missing(request, issued_id):
    issued_book = get_object_or_404(IssuedBook, id=issued_id)
    book = issued_book.book
    penalty = issued_book.calculate_missing_penalty()

    if request.method == "POST":
        payment_method = request.POST.get("payment_method")
        if not payment_method:
            messages.error(request, "⚠️ Please choose payment method.")
            return redirect("view_issued_books")

        issued_book.status = "Missing"
        issued_book.missing = True
        issued_book.final_fine = penalty
        issued_book.payment_method = payment_method
        issued_book.save()

        # reduce stock
        if book.quantity > 0:
            book.quantity -= 1
            book.save(update_fields=["quantity"])

        messages.success(
            request,
            f"📕 {book.name} marked as Missing. Penalty ₹{penalty} paid via {payment_method}."
        )
        return redirect("view_issued_books")

    # if GET, just bounce back (modal will handle UI)
    return redirect("view_issued_books")


# def mark_missing(request, issued_id):
#     issued_book = get_object_or_404(IssuedBook, id=issued_id)
#     book = issued_book.book

#     # calculate penalty (book price + late fine if overdue)
#     penalty = issued_book.calculate_missing_penalty()

#     if request.method == "POST":
#         payment_method = request.POST.get("payment_method")
#         if not payment_method:
#             messages.error(request, "⚠️ Please choose payment method to pay penalty.")
#             return redirect('view_issued_books')

#         # mark as missing
#         issued_book.status = "Missing"
#         issued_book.missing = True
#         issued_book.final_fine = penalty
#         issued_book.payment_method = payment_method
#         issued_book.save()

#         # reduce book stock
#         if book.quantity > 0:
#             book.quantity -= 1
#             book.save()

#         messages.success(request, f"📕 {book.name} marked as missing. Penalty ₹{penalty} paid.")
#         return redirect('view_issued_books')

#     return render(request, "library/view_issued_books.html", {"issued_book": issued_book, "penalty": penalty})


@staff_member_required
def cancel_reservation(request, issued_book_id):
    issued_book = get_object_or_404(IssuedBook, id=issued_book_id)

    canceled_by = 'admin' if request.user.is_staff else 'user'

    # Cancel the current issued record
    issued_book.canceled = True
    issued_book.canceled_by = canceled_by
    issued_book.status = 'Returned'
    print("Cancel Flag Set:", issued_book.canceled)
    issued_book.save()

    # Check how many reserved users are already active
    reserved_limit = issued_book.book.quantity
    current_reserved_count = IssuedBook.objects.filter(
        book=issued_book.book,
        status='Reserved',
        canceled=False,
        returned=False
    ).count()

    # Promote if slots available
    if current_reserved_count < reserved_limit:
        next_waiting = IssuedBook.objects.filter(
            book=issued_book.book,
            status='Waiting',
            canceled=False,
            returned=False
        ).order_by('requested_date').first()

        if next_waiting:
            next_waiting.status = 'Reserved'
            next_waiting.save()

            # Send mail to promoted user
            if next_waiting.student and next_waiting.student.email:
                try:
                    send_mail(
                        subject='Your Book Reservation is Confirmed!',
                        message=f'Hello {next_waiting.student.username},\n\nYou have been promoted to Reserved status for "{next_waiting.book.name}". Please check your dashboard.',
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        recipient_list=[next_waiting.student.email],
                    )
                except Exception as e:
                    print("Email error (promoted):", e)

    # Notify canceled user
    if issued_book.student and issued_book.student.email:
        try:
            send_mail(
                subject='Your Reservation has been Cancelled',
                message=f'Hello {issued_book.student.username},\n\nYour reservation for "{issued_book.book.name}" has been cancelled by the {canceled_by}.',
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[issued_book.student.email],
            )
        except Exception as e:
            print("Email error (cancelled):", e)

    messages.success(request, f'Reservation for "{issued_book.book.name}" has been cancelled and next user (if any) promoted.')
    return redirect('view_issued_books')

@staff_member_required
def approve_return(request, issued_book_id):
    issued_book = get_object_or_404(IssuedBook, id=issued_book_id)
    issued_book.returned = True
    issued_book.returned_date = timezone.now().date()  # ✅ Set returned date
    issued_book.save()
    messages.success(request, f"Return approved for {issued_book.book.name}.")
    return redirect('view_issued_books')

@login_required
def view_canceled_orders(request):
    if request.user.is_staff:
        canceled_books = IssuedBook.objects.filter(canceled=True).order_by('-requested_date')
    else:
        canceled_books = IssuedBook.objects.filter(student=request.user, canceled=True).order_by('-requested_date')

    return render(request, 'library/view_canceled_orders.html', {'canceled_books': canceled_books})


@staff_member_required
@require_POST
def update_issue_status(request, pk):
    record = get_object_or_404(IssuedBook, pk=pk)
    book = record.book

    if record.status == 'Borrowed':
        record.returned = True  # explicitly mark as returned
        record.returned_date = now()
        record.status = 'Returned'
        record.save()
        promote_users(book)

        messages.success(request, f"Issue status updated for '{book.name}'.")
    else:
        messages.warning(request, "Only Borrowed books can be marked as Returned or Not Returned.")

    messages.success(request, f'Reservation for "{record.book.name}" has been cancelled and next user updated.')
    return redirect('view_issued_books')

def delete_issue_record(request, pk):
    issue = get_object_or_404(IssuedBook, pk=pk)
    issue.delete()
    messages.success(request, "Issue record deleted.")
    return redirect('view_issued_books')

def assign_waiting_books():
    for queue in BookRequestQueue.objects.filter(fulfilled=False).order_by('requested_at'):
        if queue.book.quantity > 0:
            IssuedBook.objects.create(student=queue.student, book=queue.book)
            queue.book.quantity -= 1
            queue.book.save()
            queue.fulfilled = True
            queue.save()



# def genre_books_view(request, genre_id):
#     genre = Genre.objects.get(pk=genre_id)
#     books = Book.objects.filter(genre=genre)
#     return render(request, 'library/genre_books.html', {'genre': genre, 'books': books})

from django.db.models import Min


def genre_books_view(request, genre_id):
    genre = Genre.objects.get(pk=genre_id)

    # Pick the lowest id for each unique book name within this genre
    unique_books = (
        Book.objects.filter(genre=genre)
        .values('name', 'author')                # group by book name
        .annotate(first_id=Min('id'))  # pick the first occurrence
    )

    # Now fetch those book objects
    books = Book.objects.filter(id__in=[b['first_id'] for b in unique_books])

    return render(request, 'library/genre_books.html', {
        'genre': genre,
        'books': books
    })


# add genre in the index page
@staff_member_required
def add_genre(request):
    if request.method == 'POST':
        form = GenreForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Genre added successfully!")
            return redirect('view_genres')  # or any other view
    else:
        form = GenreForm()
    return render(request, 'library/add_genre.html', {'form': form})



# @staff_member_required
# def view_genres(request):
#     genre_list = Genre.objects.all().order_by('-id')

#     query = request.GET.get('q', '')
#     if query:
#         genre_list = genre_list.filter(
#             Q(name__icontains=query)
#         )

#     try:
#         entries_per_page = int(request.GET.get('entries', 10))
#     except (ValueError, TypeError):
#         entries_per_page = 10

#     paginator = Paginator(genre_list, entries_per_page)
#     page_number = request.GET.get('page')
#     page_obj = paginator.get_page(page_number)

#     return render(request, 'library/view_genres.html', {
#         'genres': page_obj,
#         'query': query,
#         'entries_per_page': entries_per_page,
#     })


import openpyxl
from django.http import HttpResponse
from django.db.models import Q
from django.core.paginator import Paginator
from django.contrib.admin.views.decorators import staff_member_required
from .models import Genre


@staff_member_required
def view_genres(request):
    genre_list = Genre.objects.all().order_by('-id')

    query = request.GET.get('q', '')
    if query:
        genre_list = genre_list.filter(
            Q(name__icontains=query)
        )

    # ✅ Handle export
    if "export" in request.GET:
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Genres"

        # Header row
        ws.append(["ID", "Name"])

        # Data rows
        for genre in genre_list:
            ws.append([genre.id, genre.name])

        # Response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="genres.xlsx"'
        wb.save(response)
        return response

    # ✅ Pagination
    try:
        entries_per_page = int(request.GET.get('entries', 10))
    except (ValueError, TypeError):
        entries_per_page = 10

    paginator = Paginator(genre_list, entries_per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'library/view_genres.html', {
        'genres': page_obj,
        'query': query,
        'entries_per_page': entries_per_page,
    })


@staff_member_required
def edit_genre(request, genre_id):
    genre = get_object_or_404(Genre, id=genre_id)
    if request.method == 'POST':
        form = GenreForm(request.POST, request.FILES, instance=genre)
        if form.is_valid():
            form.save()
            return redirect('view_genres')
    else:
        form = GenreForm(instance=genre)
    return render(request, 'library/edit_genre.html', {'form': form, 'genre': genre})

@staff_member_required
def delete_genre(request, genre_id):
    genre = get_object_or_404(Genre, id=genre_id)
    genre.delete()
    return redirect('view_genres')

def promote_users(book):
    """
    Promotes users in the queue when:
    - A book is returned
    - A reservation is cancelled
    """

    total_quantity = book.quantity

    borrowed_users = IssuedBook.objects.filter(book=book, status='Borrowed', returned=False).count()
    reserved_users = IssuedBook.objects.filter(book=book, status='Reserved').order_by('issued_date')
    waiting_users = IssuedBook.objects.filter(book=book, status='Waiting').order_by('issued_date')

    available_slots = total_quantity - borrowed_users

    # Promote Reserved → Borrowed
    for reserved in reserved_users:
        if available_slots <= 0:
            break

        reserved.status = 'Borrowed'
        reserved.save()
        available_slots -= 1

        subject = f"📚 Borrow Now: {book.name}"
        message = f"Dear {reserved.student.username},\n\nYou've been promoted to *Borrowed* status for the book '{book.name}'. Please collect it within the next 5 days."

        send_book_notification_email(
            reserved.student.email,
            subject,
            message
        )

    # Promote Waiting → Reserved
    current_reserved = IssuedBook.objects.filter(book=book, status='Reserved').count()
    max_reservations = total_quantity
    remaining_reservation_slots = max_reservations - current_reserved

    for waiting in waiting_users:
        if remaining_reservation_slots <= 0:
            break

        waiting.status = 'Reserved'
        waiting.save()
        remaining_reservation_slots -= 1

        subject = f"📖 Reserved for You: {book.name}"
        message = f"Dear {waiting.student.username},\n\nYou've been promoted to *Reserved* status for the book '{book.name}'. Please keep an eye on your dashboard to borrow it soon."

        send_book_notification_email(
            waiting.student.email,
            subject,
            message
        )


def send_expiry_reminders():
    today = now().date()
    books = IssuedBook.objects.filter(expiry_date=today, returned=False)

    for book in books:
        subject = "⚠️ Library Book Return Reminder"
        message = (
            f"Dear {book.student.username},\n\n"
            f"This is a reminder that your borrowed book '{book.name}' "
            f"is due today ({book.expiry_date}). Please return it on time to avoid fines.\n\n"
            "Regards,\nLibrary Team"
        )
        send_mail(subject, message, "pprathisha4@gmail.com", [book.student.email])

from django.http import HttpResponse
from .utils import send_expiry_reminders   # import the function you wrote earlier

def check_expiry_and_notify(request):
    send_expiry_reminders()
    return HttpResponse("Expiry reminders sent!")

# def mark_missing(request, issued_book_id):
#     issued_book = get_object_or_404(IssuedBook, id=issued_book_id)

#     if issued_book.missing:
#         return JsonResponse({"error": "This book is already marked as missing."}, status=400)

#     penalty = issued_book.book.price
#     issued_book.missing = True
#     issued_book.fine = penalty
#     issued_book.save()

#     # Update book quantity ↓
#     if issued_book.book.quantity > 0:
#         issued_book.book.quantity -= 1
#         issued_book.book.save()

#     return JsonResponse({"success": f"{issued_book.book.name} marked as missing. Penalty = ₹{penalty}"})



def holiday_list(request):
    holidays = Holiday.objects.all().order_by('date')
    form = HolidayForm()

    # Handle Add Holiday
    if request.method == "POST":
        form = HolidayForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Holiday added successfully!")
            return redirect('holiday_list')

    return render(request, "library/holiday_list.html", {"holidays": holidays, "form": form})

def edit_holiday(request, pk):
    holiday = get_object_or_404(Holiday, pk=pk)
    if request.method == "POST":
        form = HolidayForm(request.POST, instance=holiday)
        if form.is_valid():
            form.save()
            messages.success(request, "✏️ Holiday updated successfully!")
            return redirect('holiday_list')
    else:
        form = HolidayForm(instance=holiday)
    return render(request, "library/edit_holiday.html", {"form": form, "edit": True})

def delete_holiday(request, pk):
    holiday = get_object_or_404(Holiday, pk=pk)
    holiday.delete()
    messages.success(request, "🗑️ Holiday deleted successfully!")
    return redirect('holiday_list')
